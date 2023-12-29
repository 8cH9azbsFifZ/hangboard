import openpyxl
from pathlib import Path

xlsx_file = Path('exercises/doc', 'exercises.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file, data_only = True)  # evaluate formulas
sheet = wb_obj.active
#print(sheet["C2"].value)
col_names = []
for column in sheet.iter_cols(1, sheet.max_column):
    col_names.append(column[0].value)
   
#    ['Session', 'Rest-to-Start', 'Exercise', 'Counter', 'Pause', 'Reps', 'Left', 'Right']
#print(col_names)



wa = {}
wa["Workouts"] = []

wo = {}
wo["Sets"] = []

prevsession=0
for i, row in enumerate(sheet.iter_rows(values_only=True)):
    if i == 0:
        continue
    td = {}
    session =str(row[0])
    td["Rest-to-Start"] =str(row[1])
    td["Exercise"] =row[2]
    td["Counter"] =str(row[3])
    td["Pause"]=str(row[4])
    td["Reps"] =str(row[5])
    td["Left"] =row[6]
    td["Right"]=row[7]

    if prevsession != session:
        wa["Workouts"].append(wo)
        wo = {}
        wo["Sets"] = []

    wo["Sets"].append(td)

    
    #if prevsession != session:
    #    if session > 1:
    #        print ("]") # sets
    #        print ("}") # workout
    #    print ("{") # workout
    #    print ('"Name:" "Workout name",') # FIXME
    #    print ('"ID": ' + session + ",") # FIXME
    #    print ('"Sets":[')
    
    #if 
    #set = '{ "": '+resttostart+', "": "'+exercise+'","": '+counter+', "": '+pause+', "": '+reps+', "": "'+left+'", "": "'+right+'"}' 
    print (td)
#print ("]")
#print ("}")

print (wo)
print (wa)