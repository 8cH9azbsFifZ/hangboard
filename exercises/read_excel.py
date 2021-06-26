import openpyxl
from pathlib import Path

xlsx_file = Path('exercises/doc', 'exercises.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file, data_only = True)  # evaluate formulas
sheet = wb_obj.active
#print(sheet["C2"].value)
col_names = []
for column in sheet.iter_cols(1, sheet.max_column):
    col_names.append(column[0].value)
   
#    ['Session', 'Rest-to-Start', 'Exercise', 'Counter', 'Pause', 'Reps', 'Left', 'Right']
#print(col_names)



#print ("{")
#print ("Workouts:[")

data = {}

prevsession=0
for i, row in enumerate(sheet.iter_rows(values_only=True)):
    session=str(row[0])
    resttostart=str(row[1])
    exercise=row[2]
    counter=str(row[3])
    pause=str(row[4])
    reps=str(row[5])
    left=row[6]
    right=row[7]
    data["Workouts"][session] = 1
    
    #if prevsession != session:
    #    if session > 1:
    #        print ("]") # sets
    #        print ("}") # workout
    #    print ("{") # workout
    #    print ('"Name:" "Workout name",') # FIXME
    #    print ('"ID": ' + session + ",") # FIXME
    #    print ('"Sets":[')
    
    #if 
    #set = '{ "Rest-to-Start": '+resttostart+', "Exercise": "'+exercise+'","Counter": '+counter+', "Pause": '+pause+', "Reps": '+reps+', "Left": "'+left+'", "Right": "'+right+'"}' 
    print (set)
#print ("]")
#print ("}")

print (data)